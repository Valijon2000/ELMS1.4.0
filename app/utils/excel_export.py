from datetime import datetime
from flask import Response
import io


def create_students_excel(students, faculty_name=None):
    """Talabalar ro'yxatini Excel formatida yaratish"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. Iltimos, 'pip install openpyxl' buyrug'ini bajaring.")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Talabalar"
    
    # Sarlavha
    title = f"Talabalar ro'yxati"
    if faculty_name:
        title += f" - {faculty_name}"
    
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Sana
    ws['A2'] = f"Yaratilgan: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws.merge_cells('A2:H2')
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Jadval sarlavhalari
    headers = ['№', 'Talaba ID', 'To\'liq ism', 'Email', 'Telefon', 'Guruh', 'Qabul yili', 'Holat']
    header_row = 3
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Ma'lumotlar
    for row_num, student in enumerate(students, start=header_row + 1):
        ws.cell(row=row_num, column=1, value=row_num - header_row)
        ws.cell(row=row_num, column=2, value=student.student_id or '')
        ws.cell(row=row_num, column=3, value=student.full_name)
        ws.cell(row=row_num, column=4, value=student.email)
        ws.cell(row=row_num, column=5, value=student.phone or '')
        ws.cell(row=row_num, column=6, value=student.group.name if student.group else '')
        ws.cell(row=row_num, column=7, value=student.enrollment_year or '')
        ws.cell(row=row_num, column=8, value='Faol' if student.is_active else 'Nofaol')
        
        # Stil
        for col_num in range(1, 9):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Ustun kengliklarini sozlash
    column_widths = [5, 15, 30, 25, 15, 15, 12, 12]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Excel faylni qaytarish
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def create_schedule_excel(schedules, group_name=None, faculty_name=None):
    """Dars jadvalini Excel formatida yaratish"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. Iltimos, 'pip install openpyxl' buyrug'ini bajaring.")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Dars jadvali"
    
    # Sarlavha
    title = "Dars jadvali"
    if group_name:
        title += f" - {group_name}"
    elif faculty_name:
        title += f" - {faculty_name}"
    
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Sana
    ws['A2'] = f"Yaratilgan: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws.merge_cells('A2:G2')
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Jadval sarlavhalari
    headers = ['№', 'Hafta kuni', 'Vaqt', 'Fan', 'O\'qituvchi', 'Xona', 'Dars turi']
    header_row = 3
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Hafta kunlari
    week_days = {
        0: 'Dushanba',
        1: 'Seshanba',
        2: 'Chorshanba',
        3: 'Payshanba',
        4: 'Juma',
        5: 'Shanba'
    }
    
    # Ma'lumotlar
    for row_num, schedule in enumerate(schedules, start=header_row + 1):
        ws.cell(row=row_num, column=1, value=row_num - header_row)
        ws.cell(row=row_num, column=2, value=week_days.get(schedule.day_of_week, ''))
        ws.cell(row=row_num, column=3, value=f"{schedule.start_time} - {schedule.end_time}")
        ws.cell(row=row_num, column=4, value=schedule.subject.name if schedule.subject else '')
        ws.cell(row=row_num, column=5, value=schedule.teacher.full_name if schedule.teacher else '')
        ws.cell(row=row_num, column=6, value=schedule.link or '')
        ws.cell(row=row_num, column=7, value=schedule.lesson_type or '')
        
        # Stil
        for col_num in range(1, 8):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Ustun kengliklarini sozlash
    column_widths = [5, 15, 20, 30, 25, 15, 15]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Excel faylni qaytarish
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def create_contracts_excel(payments, course_year=None):
    """Kontrakt ma'lumotlarini Excel formatida yaratish (kurs bo'yicha)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. Iltimos, 'pip install openpyxl' buyrug'ini bajaring.")
    
    wb = Workbook()
    
    # Kurs bo'yicha guruhlash
    from collections import defaultdict
    payments_by_course = defaultdict(list)
    
    for payment in payments:
        if payment.student and payment.student.group:
            course = payment.student.group.course_year
            payments_by_course[course].append(payment)
    
    # Har bir kurs uchun alohida worksheet
    for course in sorted(payments_by_course.keys()):
        ws = wb.create_sheet(title=f"{course}-kurs")
        
        # Sarlavha
        title = f"{course}-kurs talabalar kontrakt ma'lumotlari"
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Sana
        ws['A2'] = f"Yaratilgan: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws.merge_cells('A2:H2')
        ws['A2'].font = Font(size=10, italic=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Jadval sarlavhalari
        headers = ['№', 'Talaba ID', 'To\'liq ism', 'Guruh', 'Kontrakt miqdori', 'To\'lagan', 'Qolgan', 'Foiz']
        header_row = 3
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Ma'lumotlar
        course_payments = payments_by_course[course]
        total_contract = 0
        total_paid = 0
        
        for row_num, payment in enumerate(course_payments, start=header_row + 1):
            student = payment.student
            contract = float(payment.contract_amount)
            paid = float(payment.paid_amount)
            remaining = contract - paid
            percentage = payment.get_payment_percentage()
            
            total_contract += contract
            total_paid += paid
            
            ws.cell(row=row_num, column=1, value=row_num - header_row)
            ws.cell(row=row_num, column=2, value=student.student_id or '')
            ws.cell(row=row_num, column=3, value=student.full_name)
            ws.cell(row=row_num, column=4, value=student.group.name if student.group else '')
            ws.cell(row=row_num, column=5, value=contract)
            ws.cell(row=row_num, column=6, value=paid)
            ws.cell(row=row_num, column=7, value=remaining)
            ws.cell(row=row_num, column=8, value=f"{percentage}%")
            
            # Stil
            for col_num in range(1, 9):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                # Foiz bo'yicha rang
                if col_num == 8:
                    if percentage == 100:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(bold=True, color="006100")
                    elif percentage >= 75:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        cell.font = Font(bold=True, color="9C6500")
                    elif percentage >= 50:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(bold=True, color="9C0006")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(bold=True, color="9C0006")
        
        # Jami qator
        summary_row = header_row + len(course_payments) + 2
        ws.cell(row=summary_row, column=3, value="JAMI:")
        ws.cell(row=summary_row, column=3).font = Font(bold=True, size=12)
        ws.cell(row=summary_row, column=5, value=total_contract)
        ws.cell(row=summary_row, column=5).font = Font(bold=True, size=12)
        ws.cell(row=summary_row, column=6, value=total_paid)
        ws.cell(row=summary_row, column=6).font = Font(bold=True, size=12)
        ws.cell(row=summary_row, column=7, value=total_contract - total_paid)
        ws.cell(row=summary_row, column=7).font = Font(bold=True, size=12)
        
        # Ustun kengliklarini sozlash
        column_widths = [5, 15, 30, 15, 18, 18, 18, 10]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Bosh worksheet'ni o'chirish
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Excel faylni qaytarish
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def create_group_grades_excel(subject, group, student_rows):
    """Guruh bo'yicha baholarni Excel formatida yaratish"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. Iltimos, 'pip install openpyxl' buyrug'ini bajaring.")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Baholar"
    
    # Sarlavha
    title = f"{subject.name} - {group.name} guruh baholari"
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Sana
    ws['A2'] = f"Yaratilgan: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws.merge_cells('A2:H2')
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    headers = ['№', 'Talaba ID', "To'liq ism", 'Guruh', 'Fan', 'Umumiy ball', 'Maks ball', 'Foiz', 'Baho']
    header_row = 3
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    for row_num, row in enumerate(student_rows, start=header_row + 1):
        student = row['student']
        percent = row['percent']
        grade = row['grade']
        
        ws.cell(row=row_num, column=1, value=row_num - header_row)
        ws.cell(row=row_num, column=2, value=student.student_id or '')
        ws.cell(row=row_num, column=3, value=student.full_name)
        ws.cell(row=row_num, column=4, value=group.name)
        ws.cell(row=row_num, column=5, value=subject.name)
        ws.cell(row=row_num, column=6, value=row['total'])
        ws.cell(row=row_num, column=7, value=row['max_total'])
        ws.cell(row=row_num, column=8, value=f"{percent}%")
        ws.cell(row=row_num, column=9, value=f"{grade.letter} - {grade.name}" if grade else "Baholanmagan")
        
        for col_num in range(1, 10):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Foiz ranglari
        percent_cell = ws.cell(row=row_num, column=8)
        if percent >= 86:
            percent_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            percent_cell.font = Font(bold=True, color="006100")
        elif percent >= 71:
            percent_cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
            percent_cell.font = Font(bold=True, color="006100")
        elif percent >= 56:
            percent_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            percent_cell.font = Font(bold=True, color="9C6500")
        elif percent >= 41:
            percent_cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            percent_cell.font = Font(bold=True, color="9C6500")
        else:
            percent_cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
            percent_cell.font = Font(bold=True, color="9C0006")
    
    # Ustun kengliklari
    widths = [5, 12, 30, 14, 24, 12, 12, 10, 16]
    for col_num, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_sample_contracts_excel():
    """Kontrakt import uchun namuna Excel fayl yaratish"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. Iltimos, 'pip install openpyxl' buyrug'ini bajaring.")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Namuna"
    
    # Sarlavha
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = "Kontrakt ma'lumotlarini import qilish uchun namuna"
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Sarlavha qatori
    headers = ['Talaba_id', 'Ismi', 'Kontrakt miqdori', 'To\'lagani']
    header_row = 2
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Namuna ma'lumotlar
    sample_data = [
        ['376100000000000', 'Tursunov Avazbek', '12652200', '6520020'],
        ['376100000000001', 'Karimova Malika', '12652200', '12652200'],
        ['376100000000002', 'Rahimov Dilshod', '12652200', '6326100'],
        ['376100000000003', 'Aliyeva Nodira', '12652200', '0'],
    ]
    
    for row_num, row_data in enumerate(sample_data, start=header_row + 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Eslatma qatori
    note_row = header_row + len(sample_data) + 2
    ws.merge_cells(f'A{note_row}:D{note_row}')
    note_cell = ws[f'A{note_row}']
    note_cell.value = "ESLATMA: Talaba_id yoki Ismi orqali talaba topiladi. Kontrakt miqdori majburiy."
    note_cell.font = Font(size=10, italic=True, color="666666")
    note_cell.alignment = Alignment(horizontal='left', vertical='center')
    note_cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
    
    # Ustun kengliklarini sozlash
    column_widths = [20, 30, 18, 18]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Excel faylni qaytarish
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

