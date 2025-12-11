from flask import flash
import io
import re


def import_students_from_excel(file, faculty_id=None):
    """Excel fayldan talabalar ro'yxatini import qilish"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. Iltimos, 'pip install openpyxl' buyrug'ini bajaring.")
    
    from app.models import User, Group
    from app import db
    
    try:
        # Excel faylni o'qish
        wb = load_workbook(filename=io.BytesIO(file.read()))
        ws = wb.active
        
        imported_count = 0
        errors = []
        
        # Sarlavha qatorini topish (3-qator yoki 1-qator)
        header_row = None
        for row_idx in range(1, min(5, ws.max_row + 1)):
            row_values = [str(cell.value).strip() if cell.value else '' for cell in ws[row_idx]]
            # Sarlavhalarni tekshirish
            if any('ism' in val.lower() or 'name' in val.lower() for val in row_values):
                header_row = row_idx
                break
        
        if header_row is None:
            header_row = 1  # Agar topilmasa, 1-qatordan boshlash
        
        # Ustunlar indekslarini topish
        headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[header_row]]
        
        # Ustunlar indekslarini aniqlash
        col_indices = {}
        for idx, header in enumerate(headers, 1):
            header_lower = header.lower()
            if 'id' in header_lower and 'talaba' in header_lower:
                col_indices['student_id'] = idx
            elif 'ism' in header_lower or 'name' in header_lower or 'to\'liq' in header_lower:
                col_indices['full_name'] = idx
            elif 'email' in header_lower:
                col_indices['email'] = idx
            elif 'telefon' in header_lower or 'phone' in header_lower:
                col_indices['phone'] = idx
            elif 'guruh' in header_lower or 'group' in header_lower:
                col_indices['group'] = idx
            elif 'qabul' in header_lower or 'enrollment' in header_lower or 'yil' in header_lower:
                col_indices['enrollment_year'] = idx
        
        # Ma'lumotlarni o'qish
        for row_idx in range(header_row + 1, ws.max_row + 1):
            row = ws[row_idx]
            
            # Bo'sh qatorni o'tkazib yuborish
            if not any(cell.value for cell in row):
                continue
            
            try:
                # Ma'lumotlarni olish
                full_name = str(row[col_indices.get('full_name', 1) - 1].value or '').strip()
                email = str(row[col_indices.get('email', 1) - 1].value or '').strip()
                
                # Minimal tekshiruvlar
                if not full_name or not email:
                    continue
                
                if '@' not in email:
                    errors.append(f"Qator {row_idx}: Noto'g'ri email format - {email}")
                    continue
                
                # Email takrorlanmasligini tekshirish
                if User.query.filter_by(email=email).first():
                    errors.append(f"Qator {row_idx}: Email allaqachon mavjud - {email}")
                    continue
                
                # Talaba ID
                student_id = None
                if 'student_id' in col_indices:
                    student_id_val = row[col_indices['student_id'] - 1].value
                    if student_id_val:
                        student_id = str(student_id_val).strip()
                        # Talaba ID takrorlanmasligini tekshirish
                        if student_id and User.query.filter_by(student_id=student_id).first():
                            errors.append(f"Qator {row_idx}: Talaba ID allaqachon mavjud - {student_id}")
                            continue
                
                # Telefon
                phone = None
                if 'phone' in col_indices:
                    phone_val = row[col_indices['phone'] - 1].value
                    if phone_val:
                        phone = str(phone_val).strip()
                
                # Guruh
                group_id = None
                if 'group' in col_indices:
                    group_name = str(row[col_indices['group'] - 1].value or '').strip()
                    if group_name:
                        # Guruh nomini topish
                        group = Group.query.filter_by(name=group_name.upper()).first()
                        if group:
                            # Fakultet tekshiruvi
                            if faculty_id and group.faculty_id != faculty_id:
                                errors.append(f"Qator {row_idx}: Guruh boshqa fakultetga tegishli - {group_name}")
                                continue
                            group_id = group.id
                        else:
                            errors.append(f"Qator {row_idx}: Guruh topilmadi - {group_name}")
                            continue
                
                # Qabul yili
                enrollment_year = None
                if 'enrollment_year' in col_indices:
                    year_val = row[col_indices['enrollment_year'] - 1].value
                    if year_val:
                        try:
                            enrollment_year = int(year_val)
                        except (ValueError, TypeError):
                            pass
                
                # Talaba yaratish
                student = User(
                    email=email,
                    full_name=full_name,
                    role='student',
                    student_id=student_id,
                    group_id=group_id,
                    enrollment_year=enrollment_year,
                    phone=phone
                )
                
                # Standart parol (foydalanuvchi keyinchalik o'zgartirishi mumkin)
                student.set_password('student123')
                
                db.session.add(student)
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Qator {row_idx}: Xatolik - {str(e)}")
                continue
        
        # Ma'lumotlarni saqlash
        db.session.commit()
        
        return {
            'success': True,
            'imported': imported_count,
            'errors': errors
        }
        
    except Exception as e:
        return {
            'success': False,
            'imported': 0,
            'errors': [f"Fayl o'qishda xatolik: {str(e)}"]
        }


def import_payments_from_excel(file):
    """Excel fayldan to'lov ma'lumotlarini import qilish"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. Iltimos, 'pip install openpyxl' buyrug'ini bajaring.")
    
    from app.models import User, StudentPayment
    from app import db
    
    def parse_amount(value):
        """Summani raqamga aylantirish (bo'shliqlarni olib tashlash)"""
        if value is None:
            return 0
        if isinstance(value, (int, float)):
            return float(value)
        # Matndan raqamni ajratish
        value_str = str(value).replace(' ', '').replace(',', '')
        try:
            return float(value_str)
        except (ValueError, TypeError):
            return 0
    
    try:
        # Excel faylni o'qish
        wb = load_workbook(filename=io.BytesIO(file.read()))
        ws = wb.active
        
        imported_count = 0
        updated_count = 0
        errors = []
        
        # Sarlavha qatorini topish
        header_row = None
        for row_idx in range(1, min(5, ws.max_row + 1)):
            row_values = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[row_idx]]
            # Sarlavhalarni tekshirish
            if any('talaba' in val and 'id' in val for val in row_values) or \
               any('ism' in val or 'name' in val for val in row_values):
                header_row = row_idx
                break
        
        if header_row is None:
            header_row = 1  # Agar topilmasa, 1-qatordan boshlash
        
        # Ustunlar indekslarini topish
        headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[header_row]]
        
        # Ustunlar indekslarini aniqlash
        col_indices = {}
        for idx, header in enumerate(headers, 1):
            header_lower = header.lower()
            if 'talaba' in header_lower and 'id' in header_lower:
                col_indices['student_id'] = idx
            elif 'ism' in header_lower or 'name' in header_lower or 'to\'liq' in header_lower:
                col_indices['full_name'] = idx
            elif 'kontrakt' in header_lower or 'contract' in header_lower or 'miqdori' in header_lower:
                col_indices['contract_amount'] = idx
            elif 'to\'lagan' in header_lower or 'paid' in header_lower or 'tolagan' in header_lower:
                col_indices['paid_amount'] = idx
        
        # Ma'lumotlarni o'qish
        for row_idx in range(header_row + 1, ws.max_row + 1):
            row = ws[row_idx]
            
            # Bo'sh qatorni o'tkazib yuborish
            if not any(cell.value for cell in row):
                continue
            
            try:
                # Talaba ID yoki ism orqali talabani topish
                student = None
                
                # Talaba ID orqali
                if 'student_id' in col_indices:
                    student_id_val = row[col_indices['student_id'] - 1].value
                    if student_id_val:
                        student_id_str = str(student_id_val).strip()
                        student = User.query.filter_by(student_id=student_id_str, role='student').first()
                
                # Ism orqali (agar ID topilmasa)
                if not student and 'full_name' in col_indices:
                    full_name = str(row[col_indices['full_name'] - 1].value or '').strip()
                    if full_name:
                        student = User.query.filter_by(full_name=full_name, role='student').first()
                
                if not student:
                    name_val = row[col_indices.get('full_name', 1) - 1].value if 'full_name' in col_indices else 'Noma\'lum'
                    errors.append(f"Qator {row_idx}: Talaba topilmadi - {name_val}")
                    continue
                
                # Kontrakt miqdori
                contract_amount = 0
                if 'contract_amount' in col_indices:
                    contract_val = row[col_indices['contract_amount'] - 1].value
                    contract_amount = parse_amount(contract_val)
                
                if contract_amount <= 0:
                    errors.append(f"Qator {row_idx}: Kontrakt miqdori noto'g'ri - {student.full_name}")
                    continue
                
                # To'lagan summa
                paid_amount = 0
                if 'paid_amount' in col_indices:
                    paid_val = row[col_indices['paid_amount'] - 1].value
                    paid_amount = parse_amount(paid_val)
                
                # Mavjud yozuvni topish yoki yangi yaratish
                payment = StudentPayment.query.filter_by(student_id=student.id).first()
                
                if payment:
                    # Yangilash
                    payment.contract_amount = contract_amount
                    payment.paid_amount = paid_amount
                    payment.updated_at = datetime.utcnow()
                    updated_count += 1
                else:
                    # Yangi yaratish
                    payment = StudentPayment(
                        student_id=student.id,
                        contract_amount=contract_amount,
                        paid_amount=paid_amount,
                        academic_year='2024-2025'  # Default
                    )
                    db.session.add(payment)
                    imported_count += 1
                
            except Exception as e:
                errors.append(f"Qator {row_idx}: Xatolik - {str(e)}")
                continue
        
        # Ma'lumotlarni saqlash
        db.session.commit()
        
        return {
            'success': True,
            'imported': imported_count,
            'updated': updated_count,
            'errors': errors
        }
        
    except Exception as e:
        return {
            'success': False,
            'imported': 0,
            'updated': 0,
            'errors': [f"Fayl o'qishda xatolik: {str(e)}"]
        }

